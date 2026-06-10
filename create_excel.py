import csv
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import os

csv_file = 'Project_Progress_Report.csv'
xlsx_file = 'Project_Progress_Report.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "Progress Report"

# Read CSV data
data = []
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        data.append(row)

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Write data to worksheet
for r_idx, row in enumerate(data, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Convert percentage to number
        if r_idx > 1 and c_idx == 5:
            try:
                cell.value = int(value) / 100.0
                cell.number_format = '0%'
            except ValueError:
                pass
                
        # Apply borders
        cell.border = border_thin
        
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Define Data Validation for 'Status' (Column D)
# Options: Not Started, In Progress, Completed
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D2:D{len(data)}")

# Add Conditional Formatting for 'Status'
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
yellow_font = Font(color="9C5700")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006")

ws.conditional_formatting.add(f'D2:D{len(data)}', CellIsRule(operator='equal', formula=['"Completed"'], stopIfTrue=True, fill=green_fill, font=green_font))
ws.conditional_formatting.add(f'D2:D{len(data)}', CellIsRule(operator='equal', formula=['"In Progress"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
ws.conditional_formatting.add(f'D2:D{len(data)}', CellIsRule(operator='equal', formula=['"Not Started"'], stopIfTrue=True, fill=red_fill, font=red_font))

# Add Conditional Formatting for 'Percentage' (Column E) using a color scale
color_scale_rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                                  mid_type='num', mid_value=0.5, mid_color='FFEB84',
                                  end_type='num', end_value=1, end_color='63BE7B')
ws.conditional_formatting.add(f'E2:E{len(data)}', color_scale_rule)

# Define Data Validation for 'Complexity' (Column F) and 'Priority' (Column G)
dv_complexity = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
ws.add_data_validation(dv_complexity)
dv_complexity.add(f"F2:F{len(data)}")
dv_complexity.add(f"G2:G{len(data)}")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save(xlsx_file)
print("Excel file created successfully!")

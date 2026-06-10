import csv
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

xlsx_file = 'Project_Plan_25_Days.xlsx'

# Data for the plan
data = [
    ["Module", "Task", "Status", "Completion Duration (Days)"],
    ["1. Web APIs", "Database Design & Architecture Setup", "Completed", 2],
    ["1. Web APIs", "Core & Shared Services", "Completed", 2],
    ["1. Web APIs", "Admin & Keeper APIs", "Completed", 2],
    ["1. Web APIs", "User & Journey Routing APIs", "In Progress", 2],
    ["2. Admin Website", "UI/UX Setup & Theming", "Completed", 1],
    ["2. Admin Website", "Data Grids & Management Screens", "Completed", 2],
    ["2. Admin Website", "Authentication & Analytics Dashboard", "In Progress", 1],
    ["3. Keeper Website", "Auth & Profile/Shop Onboarding", "Completed", 1],
    ["3. Keeper Website", "Offer Creation & Management Forms", "Completed", 2],
    ["3. Keeper Website", "Dashboard Analytics & Reviews", "In Progress", 1],
    ["4. User Website", "Core Setup, Theming & Auth", "Completed", 2],
    ["4. User Website", "Maps, Geocoding & Discover Features", "In Progress", 3],
    ["4. User Website", "Journey/Trip Planning & Navigation", "In Progress", 2],
    ["4. User Website", "User Profiles, History & Feed", "Completed", 1]
]

wb = Workbook()
ws = wb.active
ws.title = "Project Plan"

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Write data to worksheet
for r_idx, row in enumerate(data, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply borders
        cell.border = border_thin
        
        # Alignment for numbers
        if c_idx == 4 and r_idx > 1:
            cell.alignment = Alignment(horizontal="center")
            
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Define Data Validation for 'Status' (Column C)
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Completed"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"C2:C{len(data)}")

# Add Conditional Formatting for 'Status' (Python openpyxl conditional formatting setup)
from openpyxl.formatting.rule import CellIsRule
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
yellow_font = Font(color="9C5700")

ws.conditional_formatting.add(f'C2:C{len(data)}', CellIsRule(operator='equal', formula=['"Completed"'], stopIfTrue=True, fill=green_fill, font=green_font))
ws.conditional_formatting.add(f'C2:C{len(data)}', CellIsRule(operator='equal', formula=['"In Progress"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))

# Add Total Row
total_row_idx = len(data) + 1
ws.cell(row=total_row_idx, column=3, value="Total Days:").font = Font(bold=True)
total_cell = ws.cell(row=total_row_idx, column=4, value=sum([row[3] for row in data[1:]]))
total_cell.font = Font(bold=True)
total_cell.alignment = Alignment(horizontal="center")
total_cell.border = border_thin

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    ws.column_dimensions[column].width = max_length + 3

wb.save(xlsx_file)
print("Excel file created successfully!")
